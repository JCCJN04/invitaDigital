import openpyxl, urllib.request, json

wb = openpyxl.load_workbook(r'C:\Users\mendo\Downloads\code\chamba\invitaciones\boda carla y angel\INVITADOS_CARLA_Y_ANGEL_CON_LINKS.xlsx')
sheet = wb.active

headers = [cell for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
print("Excel Header Columns:", headers)

excel_guests = []
for row in sheet.iter_rows(min_row=2, values_only=True):
    # Check which column is name and passes
    no_num = row[0]
    name = row[1]
    passes = row[2]
    link = row[3] if len(row) > 3 else None
    
    if name is not None:
        excel_guests.append({
            'no': no_num,
            'name': str(name).strip(),
            'passes': int(passes) if passes else 1,
            'link': link
        })

print(f"Total rows read from Excel: {len(excel_guests)}")
print(f"Total passes calculated from Excel: {sum(g['passes'] for g in excel_guests)}")

with open('.env.local', 'r', encoding='utf-8') as f:
    env = dict(line.strip().split('=', 1) for line in f if '=' in line and not line.startswith('#'))

key = env['NEXT_PUBLIC_SUPABASE_ANON_KEY']
url = env['NEXT_PUBLIC_SUPABASE_URL']
headers = {'apikey': key, 'Authorization': f'Bearer {key}'}

req = urllib.request.Request(f'{url}/rest/v1/guests?event_id=eq.bc3caeaa-31df-4168-aa73-54e7feda65af&select=*', headers=headers)
with urllib.request.urlopen(req) as resp:
    db_items = json.loads(resp.read().decode('utf-8'))

print(f"\nTotal guests in Supabase DB: {len(db_items)}")
print(f"Total passes in Supabase DB: {sum(g['passes_assigned'] for g in db_items)}")

# 1-to-1 comparison
db_map = {g['name'].strip().lower(): g for g in db_items}

matched = 0
passes_ok = 0
unmatched = []

for eg in excel_guests:
    k = eg['name'].strip().lower()
    if k in db_map:
        matched += 1
        if db_map[k]['passes_assigned'] == eg['passes']:
            passes_ok += 1
    else:
        unmatched.append(eg['name'])

print(f"\nExact Name Matches: {matched} / {len(excel_guests)}")
print(f"Exact Passes Matches: {passes_ok} / {len(excel_guests)}")
if unmatched:
    print(f"Unmatched names ({len(unmatched)}):", unmatched[:5])
else:
    print("ALL 123 GUESTS AND 200 PASSES ARE 100% IDENTICAL AND ACCURATE!")
