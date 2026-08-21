import openpyxl
import json
import re
import unicodedata
import secrets
import os

EXCEL_PATH = r"C:\Users\mendo\Downloads\code\chamba\invitaciones\boda carla y angel\Template Guest List CARLA Y ANGEL (1).xlsx"
EVENT_ID = "bc3caeaa-31df-4168-aa73-54e7feda65af"
EVENT_SLUG = "carlayangel"
BASE_URL = "https://www.invitacionesdigitalesmty.com.mx/carlayangel"

def slugify(text):
    text = unicodedata.normalize('NFKD', text).encode('ascii', 'ignore').decode('utf-8')
    text = re.sub(r'[^\w\s-]', '', text).strip().lower()
    return re.sub(r'[-\s]+', '_', text)[:25]

def clean_name(raw_name):
    if not raw_name:
        return ""
    s = " ".join(str(raw_name).strip().split())
    s = re.sub(r'\s*\.\s*', ' & ', s)
    s = re.sub(r'\s+y\s+', ' & ', s, flags=re.IGNORECASE)
    words = s.split()
    formatted_words = []
    for w in words:
        if w == '&':
            formatted_words.append('&')
        elif w.lower() in ['de', 'del', 'la', 'las', 'los', 'san', 'santa', 'y', 'e']:
            formatted_words.append(w.lower())
        else:
            formatted_words.append(w.capitalize())
    return " ".join(formatted_words)

def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    sheet = wb['Hoja1']
    
    guests = []
    seen_tokens = set()
    
    for row_idx in range(13, sheet.max_row + 1):
        name_val = sheet.cell(row=row_idx, column=3).value
        adults_val = sheet.cell(row=row_idx, column=4).value
        kids_val = sheet.cell(row=row_idx, column=5).value
        
        if not name_val or not str(name_val).strip():
            continue
            
        raw_name = str(name_val).strip()
        formatted_name = clean_name(raw_name)
        
        try:
            adults = int(adults_val) if adults_val is not None else 1
        except:
            adults = 1
            
        try:
            kids = int(kids_val) if kids_val is not None else 0
        except:
            kids = 0
            
        total_passes = adults + kids
        if total_passes <= 0:
            total_passes = 1
            
        base_token = slugify(formatted_name.split('&')[0])
        rand_suffix = secrets.token_hex(2)
        token = f"{base_token}_{rand_suffix}"
        while token in seen_tokens:
            rand_suffix = secrets.token_hex(2)
            token = f"{base_token}_{rand_suffix}"
        seen_tokens.add(token)
        
        url = f"{BASE_URL}?guest={token}"
        whatsapp_msg = (
            f"¡Hola {formatted_name}! 🎉 Nos hace muy felices invitarte a nuestra boda. "
            f"Preparamos esta invitación digital con todos los detalles de nuestro gran día y tus pases asignados ({total_passes} {'pases' if total_passes > 1 else 'pase'}). "
            f"Por favor confirma tu asistencia aquí: {url}"
        )
        
        guests.append({
            "event_id": EVENT_ID,
            "name": formatted_name,
            "token": token,
            "passes_assigned": total_passes,
            "passes_confirmed": 0,
            "rsvp_status": "pending",
            "table_assigned": None,
            "qr_code_url": None,
            "whatsapp_sent": False,
            "whatsapp_delivered": False,
            "notes": f"Adultos: {adults}, Niños: {kids}" if kids > 0 else f"Adultos: {adults}",
            "url": url,
            "whatsapp_msg": whatsapp_msg
        })
        
    print(f"Total invitados procesados del Excel: {len(guests)}")
    
    # Save clean JSON of guests
    with open("carlayangel_guests.json", "w", encoding="utf-8") as f:
        json.dump(guests, f, ensure_ascii=False, indent=2)
        
    # Generate Excel for the couple with customized WhatsApp links
    out_wb = openpyxl.Workbook()
    out_sheet = out_wb.active
    out_sheet.title = "Invitados con Links"
    
    headers = ["#", "Invitado / Familia", "Pases Asignados", "Detalle", "Link Personalizado", "Mensaje WhatsApp Listo"]
    out_sheet.append(headers)
    
    for i, g in enumerate(guests, 1):
        out_sheet.append([
            i,
            g["name"],
            g["passes_assigned"],
            g["notes"],
            g["url"],
            g["whatsapp_msg"]
        ])
        
    out_path = r"C:\Users\mendo\Downloads\code\chamba\invitaciones\boda carla y angel\INVITADOS_CARLA_Y_ANGEL_CON_LINKS.xlsx"
    out_wb.save(out_path)
    print(f"Excel listo generado con todos los links en: {out_path}")

if __name__ == "__main__":
    main()
